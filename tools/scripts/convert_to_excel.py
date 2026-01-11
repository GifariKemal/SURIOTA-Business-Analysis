"""
Convert CSV Campaign Files to Styled Excel
==========================================
Converts email campaign CSV files to professionally styled Excel workbooks
with Geist font styling and attractive formatting.

Usage:
    python convert_to_excel.py
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Fill, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# Paths
TOOLS_DIR = Path(__file__).parent.parent
OUTPUT_DIR = TOOLS_DIR / "output"
CAMPAIGNS_DIR = OUTPUT_DIR / "campaigns"
LEADS_DIR = OUTPUT_DIR / "leads"

# SURIOTA Brand Colors
COLORS = {
    "primary": "1E3A5F",      # Deep navy blue
    "secondary": "2E8B57",    # Sea green
    "accent": "4169E1",       # Royal blue
    "light_bg": "F8FAFC",     # Light gray background
    "header_bg": "1E3A5F",    # Header background
    "header_text": "FFFFFF",  # White text
    "border": "E2E8F0",       # Light border
    "alt_row": "F1F5F9",      # Alternating row color
}

# Font settings (Geist or fallback to Calibri)
FONT_NAME = "Geist"  # Will fallback to Calibri if not installed


def create_styles():
    """Create named styles for the workbook."""
    styles = {}

    # Header style
    styles["header"] = {
        "font": Font(name=FONT_NAME, size=11, bold=True, color=COLORS["header_text"]),
        "fill": PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="medium", color=COLORS["primary"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }

    # Normal cell style
    styles["normal"] = {
        "font": Font(name=FONT_NAME, size=10, color="333333"),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="thin", color=COLORS["border"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }

    # Alternating row style
    styles["alt_row"] = {
        "font": Font(name=FONT_NAME, size=10, color="333333"),
        "fill": PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid"),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="thin", color=COLORS["border"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }

    # Email link style
    styles["email"] = {
        "font": Font(name=FONT_NAME, size=10, color=COLORS["accent"], underline="single"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(
            bottom=Side(style="thin", color=COLORS["border"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }

    # Title style
    styles["title"] = {
        "font": Font(name=FONT_NAME, size=16, bold=True, color=COLORS["primary"]),
        "alignment": Alignment(horizontal="left", vertical="center")
    }

    # Subtitle style
    styles["subtitle"] = {
        "font": Font(name=FONT_NAME, size=11, italic=True, color="666666"),
        "alignment": Alignment(horizontal="left", vertical="center")
    }

    return styles


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def csv_to_excel(csv_path: Path, excel_path: Path, title: str = None):
    """Convert CSV file to styled Excel workbook."""

    # Read CSV data
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        data = list(reader)

    if not data:
        print(f"[Skip] Empty file: {csv_path}")
        return

    headers = data[0]
    rows = data[1:]

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Data"

    styles = create_styles()

    # Add title row
    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        apply_style(title_cell, styles["title"])
        ws.row_dimensions[1].height = 30

        # Add subtitle with count
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        subtitle_cell = ws.cell(row=2, column=1, value=f"Total: {len(rows)} contacts | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        apply_style(subtitle_cell, styles["subtitle"])
        ws.row_dimensions[2].height = 20

        header_row = 4
    else:
        header_row = 1

    # Column widths based on content type
    column_widths = {
        "email": 35,
        "first_name": 15,
        "last_name": 15,
        "company": 30,
        "subject": 50,
        "body": 80,
        "position": 30,
        "department": 20,
        "industry": 25,
        "domain": 25,
        "confidence": 12,
        "sources": 10,
        "linkedin": 40,
    }

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header.upper().replace("_", " "))
        apply_style(cell, styles["header"])

        # Set column width
        col_letter = get_column_letter(col)
        header_lower = header.lower()
        width = column_widths.get(header_lower, 20)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[header_row].height = 25

    # Write data rows
    for row_idx, row_data in enumerate(rows, header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Check if this is an email column
            header_name = headers[col_idx - 1].lower() if col_idx <= len(headers) else ""

            if header_name == "email" and value and "@" in value:
                style = styles["email"].copy()
                if is_alt:
                    style["fill"] = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")
                apply_style(cell, style)
                cell.hyperlink = f"mailto:{value}"
            else:
                apply_style(cell, styles["alt_row"] if is_alt else styles["normal"])

        ws.row_dimensions[row_idx].height = 20

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Add autofilter
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"

    # Save workbook
    wb.save(excel_path)
    print(f"[Created] {excel_path.name} ({len(rows)} rows)")


def main():
    print("=" * 60)
    print("  SURIOTA Excel Converter")
    print("  Converting CSV to Styled Excel (.xlsx)")
    print("=" * 60)
    print()

    # Campaign files to convert
    campaigns = [
        ("EMAIL_CAMPAIGN_COMPRO.csv", "SURIOTA Email Campaign - Company Profile"),
        ("EMAIL_CAMPAIGN_SRT_MGATE.csv", "SURIOTA Email Campaign - SRT-MGATE-1210"),
        ("EMAIL_CAMPAIGN_SURGE.csv", "SURIOTA Email Campaign - SURGE Platform"),
        ("EMAIL_CAMPAIGN_SRICARE.csv", "SURIOTA Email Campaign - SRICARE"),
        ("MAILCHIMP_EXPORT.csv", "SURIOTA Mailchimp Export"),
    ]

    # Lead files to convert
    leads = [
        ("ALL_BATAM_LEADS.csv", "SURIOTA Batam Leads - All Contacts"),
        ("LEADS_srt_mgate.csv", "SURIOTA Leads - SRT-MGATE-1210 Targets"),
        ("LEADS_surge.csv", "SURIOTA Leads - SURGE Platform Targets"),
        ("LEADS_sricare.csv", "SURIOTA Leads - SRICARE Targets"),
    ]

    print("[Converting Campaign Files]")
    for csv_name, title in campaigns:
        csv_path = CAMPAIGNS_DIR / csv_name
        if csv_path.exists():
            excel_path = CAMPAIGNS_DIR / csv_name.replace(".csv", ".xlsx")
            csv_to_excel(csv_path, excel_path, title)
        else:
            print(f"[Skip] Not found: {csv_name}")

    print()
    print("[Converting Lead Files]")
    for csv_name, title in leads:
        csv_path = LEADS_DIR / csv_name
        if csv_path.exists():
            excel_path = LEADS_DIR / csv_name.replace(".csv", ".xlsx")
            csv_to_excel(csv_path, excel_path, title)
        else:
            print(f"[Skip] Not found: {csv_name}")

    print()
    print("=" * 60)
    print("  Done! Excel files created with styling.")
    print("=" * 60)


if __name__ == "__main__":
    main()
