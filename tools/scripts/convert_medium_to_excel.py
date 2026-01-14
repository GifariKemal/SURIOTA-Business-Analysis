"""
Convert Medium Company Campaign Files to Styled Excel
======================================================
"""

import csv
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from datetime import datetime

# Paths
TOOLS_DIR = Path(__file__).parent.parent
OUTPUT_DIR = TOOLS_DIR / "output"
CAMPAIGNS_DIR = OUTPUT_DIR / "campaigns"
LEADS_DIR = OUTPUT_DIR / "leads"

# SURIOTA Brand Colors
COLORS = {
    "primary": "1E3A5F",
    "secondary": "2E8B57",
    "accent": "4169E1",
    "header_bg": "1E3A5F",
    "header_text": "FFFFFF",
    "border": "E2E8F0",
    "alt_row": "F1F5F9",
}

FONT_NAME = "Calibri"


def create_styles():
    styles = {}
    styles["header"] = {
        "font": Font(name=FONT_NAME, size=11, bold=True, color=COLORS["header_text"]),
        "fill": PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="medium", color=COLORS["primary"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }
    styles["normal"] = {
        "font": Font(name=FONT_NAME, size=10, color="333333"),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="thin", color=COLORS["border"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }
    styles["alt_row"] = {
        "font": Font(name=FONT_NAME, size=10, color="333333"),
        "fill": PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid"),
        "alignment": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="thin", color=COLORS["border"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }
    styles["email"] = {
        "font": Font(name=FONT_NAME, size=10, color=COLORS["accent"], underline="single"),
        "alignment": Alignment(horizontal="left", vertical="center"),
        "border": Border(
            bottom=Side(style="thin", color=COLORS["border"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }
    styles["title"] = {
        "font": Font(name=FONT_NAME, size=16, bold=True, color=COLORS["primary"]),
        "alignment": Alignment(horizontal="left", vertical="center")
    }
    styles["subtitle"] = {
        "font": Font(name=FONT_NAME, size=11, italic=True, color="666666"),
        "alignment": Alignment(horizontal="left", vertical="center")
    }
    return styles


def apply_style(cell, style_dict):
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def csv_to_excel(csv_path: Path, excel_path: Path, title: str = None):
    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        data = list(reader)

    if not data:
        print(f"[Skip] Empty file: {csv_path}")
        return

    headers = data[0]
    rows = data[1:]

    wb = Workbook()
    ws = wb.active
    ws.title = "Campaign Data"

    styles = create_styles()

    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        apply_style(title_cell, styles["title"])
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        subtitle_cell = ws.cell(row=2, column=1, value=f"Total: {len(rows)} contacts | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        apply_style(subtitle_cell, styles["subtitle"])
        ws.row_dimensions[2].height = 20

        header_row = 4
    else:
        header_row = 1

    column_widths = {
        "email": 35, "name": 20, "company": 40, "position": 25,
        "subject": 50, "body_id": 80, "body_en": 80, "industry": 30,
        "phone": 20, "category": 20, "city": 15, "confidence": 12, "source": 15
    }

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header.upper().replace("_", " "))
        apply_style(cell, styles["header"])
        col_letter = get_column_letter(col)
        header_lower = header.lower()
        width = column_widths.get(header_lower, 20)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[header_row].height = 25

    for row_idx, row_data in enumerate(rows, header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            header_name = headers[col_idx - 1].lower() if col_idx <= len(headers) else ""
            if header_name == "email" and value and "@" in value:
                style = styles["email"].copy()
                if is_alt:
                    style["fill"] = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")
                apply_style(cell, style)
                cell.hyperlink = f"mailto:{value}"
            else:
                apply_style(cell, styles["alt_row"] if is_alt else styles["normal"])
        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(rows)}"

    wb.save(excel_path)
    print(f"[Created] {excel_path.name} ({len(rows)} rows)")


def json_to_excel(json_path: Path, excel_path: Path, title: str = None):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not data:
        print(f"[Skip] Empty file: {json_path}")
        return

    headers = list(data[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "Leads Data"

    styles = create_styles()

    if title:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        title_cell = ws.cell(row=1, column=1, value=title)
        apply_style(title_cell, styles["title"])
        ws.row_dimensions[1].height = 30

        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
        subtitle_cell = ws.cell(row=2, column=1, value=f"Total: {len(data)} contacts | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        apply_style(subtitle_cell, styles["subtitle"])
        ws.row_dimensions[2].height = 20

        header_row = 4
    else:
        header_row = 1

    column_widths = {
        "email": 35, "name": 20, "company": 40, "position": 25,
        "industry": 35, "phone": 20, "category": 20, "city": 15,
        "website": 25, "confidence": 12, "source": 15
    }

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header.upper().replace("_", " "))
        apply_style(cell, styles["header"])
        col_letter = get_column_letter(col)
        width = column_widths.get(header.lower(), 20)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[header_row].height = 25

    for row_idx, row_data in enumerate(data, header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            if header.lower() == "email" and value and "@" in str(value):
                style = styles["email"].copy()
                if is_alt:
                    style["fill"] = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")
                apply_style(cell, style)
                cell.hyperlink = f"mailto:{value}"
            else:
                apply_style(cell, styles["alt_row"] if is_alt else styles["normal"])
        ws.row_dimensions[row_idx].height = 20

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(data)}"

    wb.save(excel_path)
    print(f"[Created] {excel_path.name} ({len(data)} rows)")


def main():
    print("=" * 60)
    print("  SURIOTA Medium Company Excel Converter")
    print("=" * 60)
    print()

    timestamp = datetime.now().strftime("%Y%m%d")

    # Convert campaign CSV
    csv_name = f"EMAIL_CAMPAIGN_MEDIUM_COMPRO_{timestamp}.csv"
    csv_path = CAMPAIGNS_DIR / csv_name
    if csv_path.exists():
        excel_path = CAMPAIGNS_DIR / csv_name.replace(".csv", ".xlsx")
        csv_to_excel(csv_path, excel_path, "SURIOTA Email Campaign - Medium Companies (EPC/Fabrication)")
    else:
        print(f"[Skip] Not found: {csv_name}")

    # Convert leads JSON
    json_name = "MEDIUM_COMPANY_BATAM_COMPRO.json"
    json_path = LEADS_DIR / json_name
    if json_path.exists():
        excel_path = LEADS_DIR / json_name.replace(".json", ".xlsx")
        json_to_excel(json_path, excel_path, "SURIOTA Leads - Medium Companies Batam")

    # Also convert Micro company files if exist
    micro_csv = f"EMAIL_CAMPAIGN_MICRO_COMPRO_{timestamp}.csv"
    micro_csv_path = CAMPAIGNS_DIR / micro_csv
    if micro_csv_path.exists():
        excel_path = CAMPAIGNS_DIR / micro_csv.replace(".csv", ".xlsx")
        csv_to_excel(micro_csv_path, excel_path, "SURIOTA Email Campaign - Micro Companies")

    micro_json = "MICRO_COMPANY_BATAM_COMPRO.json"
    micro_json_path = LEADS_DIR / micro_json
    if micro_json_path.exists():
        excel_path = LEADS_DIR / micro_json.replace(".json", ".xlsx")
        json_to_excel(micro_json_path, excel_path, "SURIOTA Leads - Micro Companies Batam")

    print()
    print("=" * 60)
    print("  Done! Excel files created.")
    print("=" * 60)


if __name__ == "__main__":
    main()
