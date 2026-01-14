"""
Filter Priority Leads - Prioritas untuk Dikirim Duluan
=======================================================
Kriteria prioritas:
1. Named contact (bukan generic email)
2. Decision maker position
3. Industry relevance dengan SURIOTA
4. High confidence email
5. Company category
"""

import json
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
TOOLS_DIR = Path(__file__).parent.parent
OUTPUT_DIR = TOOLS_DIR / "output"
LEADS_DIR = OUTPUT_DIR / "leads"
CAMPAIGNS_DIR = OUTPUT_DIR / "campaigns"

# Brand Colors
COLORS = {
    "primary": "1E3A5F",
    "header_text": "FFFFFF",
    "border": "E2E8F0",
    "alt_row": "F1F5F9",
    "priority_high": "DC2626",    # Red
    "priority_medium": "F59E0B",  # Orange
    "priority_low": "6B7280",     # Gray
}

# Priority Industries (relevant to SURIOTA IoT services)
PRIORITY_INDUSTRIES = [
    "manufacturing", "fabrication", "machining", "engineering",
    "oil", "gas", "energy", "power", "electrical",
    "automation", "plc", "scada", "industrial",
    "maritime", "marine", "shipyard", "vessel",
    "logistics", "warehouse", "supply chain",
    "water", "utilities", "ipal", "wtp",
    "building", "facility", "property",
    "epc", "contractor", "construction"
]

# Decision Maker Positions
DECISION_MAKER_POSITIONS = [
    "director", "direktur", "ceo", "cto", "coo", "cfo", "cmo",
    "president", "vice president", "vp",
    "general manager", "gm", "manager",
    "head", "chief", "lead", "senior",
    "owner", "founder", "partner",
    "procurement", "purchasing", "buyer",
    "engineering manager", "plant manager", "operations manager",
    "it manager", "technical manager", "maintenance manager"
]

# Generic/Low Priority Positions
GENERIC_POSITIONS = [
    "sales", "admin", "support", "contact", "info",
    "hr", "recruitment", "hrd", "career",
    "marketing", "customer service", "receptionist"
]

# Generic Email Patterns (low priority)
GENERIC_EMAIL_PATTERNS = [
    "info@", "contact@", "sales@", "admin@", "support@",
    "hr@", "hrd@", "recruitment@", "career@", "job@",
    "marketing@", "enquiry@", "inquiry@", "general@",
    "office@", "hello@", "mail@", "email@"
]


def calculate_priority_score(lead):
    """Calculate priority score (0-100)"""
    score = 0
    reasons = []

    email = lead.get("email", "").lower()
    name = lead.get("name", "").lower()
    position = lead.get("position", "").lower()
    industry = lead.get("industry", "").lower()
    company = lead.get("company", "").lower()
    category = lead.get("category", "")
    confidence = str(lead.get("confidence", "")).lower()

    # 1. Named Contact vs Generic Email (0-25 points)
    is_generic_email = any(pattern in email for pattern in GENERIC_EMAIL_PATTERNS)
    has_real_name = name and name not in ["unknown", "bapak/ibu", "sales team", "info team", "admin team"]

    if has_real_name and not is_generic_email:
        score += 25
        reasons.append("Named contact")
    elif has_real_name:
        score += 15
        reasons.append("Has name")
    elif not is_generic_email:
        score += 10
        reasons.append("Personal email")
    else:
        reasons.append("Generic email")

    # 2. Decision Maker Position (0-25 points)
    if position:
        is_decision_maker = any(dm in position for dm in DECISION_MAKER_POSITIONS)
        is_generic_position = any(gp in position for gp in GENERIC_POSITIONS)

        if is_decision_maker:
            score += 25
            reasons.append(f"Decision maker: {position}")
        elif not is_generic_position:
            score += 15
            reasons.append(f"Technical role: {position}")
        else:
            score += 5
            reasons.append(f"Generic role: {position}")

    # 3. Industry Relevance (0-25 points)
    if industry:
        is_priority_industry = any(pi in industry for pi in PRIORITY_INDUSTRIES)
        if is_priority_industry:
            score += 25
            reasons.append(f"Priority industry: {industry[:30]}")
        else:
            score += 10
            reasons.append(f"Other industry: {industry[:30]}")

    # 4. Email Confidence (0-15 points)
    if confidence:
        try:
            conf_val = int(confidence)
            if conf_val >= 90:
                score += 15
                reasons.append(f"High confidence: {conf_val}%")
            elif conf_val >= 70:
                score += 10
                reasons.append(f"Medium confidence: {conf_val}%")
            elif conf_val >= 50:
                score += 5
                reasons.append(f"Low confidence: {conf_val}%")
        except ValueError:
            pass

    # 5. Company Category Bonus (0-10 points)
    if category == "Big Company":
        score += 10
        reasons.append("Big company")
    elif category == "Medium Company":
        score += 8
        reasons.append("Medium company")
    else:
        score += 5
        reasons.append("Micro company")

    return score, reasons


def get_priority_level(score):
    """Get priority level based on score"""
    if score >= 70:
        return "HIGH", COLORS["priority_high"]
    elif score >= 45:
        return "MEDIUM", COLORS["priority_medium"]
    else:
        return "LOW", COLORS["priority_low"]


def load_campaigns():
    """Load campaign data"""
    timestamp = datetime.now().strftime("%Y%m%d")
    path = CAMPAIGNS_DIR / f"EMAIL_CAMPAIGN_ALL_LEADS_{timestamp}.json"

    if not path.exists():
        files = list(CAMPAIGNS_DIR.glob("EMAIL_CAMPAIGN_ALL_LEADS_*.json"))
        if files:
            path = sorted(files)[-1]
        else:
            print("[Error] No campaign file found")
            return []

    with open(path, "r", encoding="utf-8") as f:
        campaigns = json.load(f)

    print(f"[Loaded] {len(campaigns)} campaigns from {path.name}")
    return campaigns


def filter_and_prioritize(campaigns):
    """Filter and prioritize campaigns"""
    prioritized = []

    for campaign in campaigns:
        score, reasons = calculate_priority_score(campaign)
        level, color = get_priority_level(score)

        prioritized.append({
            **campaign,
            "priority_score": score,
            "priority_level": level,
            "priority_reasons": "; ".join(reasons)
        })

    # Sort by priority score (highest first)
    prioritized.sort(key=lambda x: -x["priority_score"])

    return prioritized


def save_priority_excel(campaigns, path, title):
    """Save prioritized campaigns to Excel"""
    if not campaigns:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Priority Leads"

    # Columns for Excel
    excel_columns = ["priority_level", "priority_score", "email", "name", "company",
                     "position", "industry", "category", "city", "subject", "priority_reasons"]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(excel_columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=COLORS["primary"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Stats
    high_count = sum(1 for c in campaigns if c.get("priority_level") == "HIGH")
    medium_count = sum(1 for c in campaigns if c.get("priority_level") == "MEDIUM")
    low_count = sum(1 for c in campaigns if c.get("priority_level") == "LOW")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(excel_columns))
    subtitle = f"Total: {len(campaigns)} | HIGH: {high_count} | MEDIUM: {medium_count} | LOW: {low_count} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(name="Calibri", size=11, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4

    # Column widths
    column_widths = {
        "priority_level": 12, "priority_score": 10, "email": 35, "name": 20,
        "company": 30, "position": 20, "industry": 25, "category": 15,
        "city": 12, "subject": 40, "priority_reasons": 50
    }

    # Header style
    header_fill = PatternFill(start_color=COLORS["primary"], end_color=COLORS["primary"], fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=COLORS["header_text"])
    border = Border(
        bottom=Side(style="thin", color=COLORS["border"]),
        top=Side(style="thin", color=COLORS["border"]),
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"])
    )

    # Headers
    for col, header in enumerate(excel_columns, 1):
        cell = ws.cell(row=header_row, column=col, value=header.upper().replace("_", " "))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = column_widths.get(header, 15)

    ws.row_dimensions[header_row].height = 25

    # Data rows
    for row_idx, campaign in enumerate(campaigns, header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        priority = campaign.get("priority_level", "LOW")

        for col_idx, header in enumerate(excel_columns, 1):
            value = campaign.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

            cell.font = Font(name="Calibri", size=10, color="333333")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border

            if is_alt:
                cell.fill = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")

            # Priority level color
            if header == "priority_level":
                if priority == "HIGH":
                    cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["priority_high"])
                elif priority == "MEDIUM":
                    cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["priority_medium"])
                else:
                    cell.font = Font(name="Calibri", size=10, bold=True, color=COLORS["priority_low"])

            # Email hyperlink
            if header == "email" and value and "@" in str(value):
                cell.font = Font(name="Calibri", size=10, color="4169E1", underline="single")
                cell.hyperlink = f"mailto:{value}"

        ws.row_dimensions[row_idx].height = 20

    # Freeze and filter
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(excel_columns))}{header_row + len(campaigns)}"

    wb.save(path)
    print(f"[Saved] {path.name}")


def save_priority_csv(campaigns, path):
    """Save prioritized campaigns to CSV"""
    if not campaigns:
        return

    # Select key columns
    columns = ["priority_level", "priority_score", "email", "name", "company",
               "position", "industry", "category", "city", "subject", "priority_reasons"]

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(campaigns)

    print(f"[Saved] {path.name}")


def main():
    print("=" * 60)
    print("  SURIOTA Priority Lead Filter")
    print("  Filter & Prioritize for Email Campaign")
    print("=" * 60)
    print()

    # Load campaigns
    campaigns = load_campaigns()
    if not campaigns:
        return

    # Filter and prioritize
    print("\n[Analyzing and prioritizing leads...]")
    prioritized = filter_and_prioritize(campaigns)

    # Count by priority
    high = [c for c in prioritized if c["priority_level"] == "HIGH"]
    medium = [c for c in prioritized if c["priority_level"] == "MEDIUM"]
    low = [c for c in prioritized if c["priority_level"] == "LOW"]

    print(f"\n[Priority Distribution]")
    print(f"  HIGH:   {len(high)} leads (score >= 70)")
    print(f"  MEDIUM: {len(medium)} leads (score 45-69)")
    print(f"  LOW:    {len(low)} leads (score < 45)")

    # Save outputs
    print("\n[Saving files...]")
    timestamp = datetime.now().strftime("%Y%m%d")

    # Save ALL prioritized (sorted by score)
    all_path = CAMPAIGNS_DIR / f"PRIORITY_ALL_LEADS_{timestamp}.xlsx"
    save_priority_excel(prioritized, all_path, "SURIOTA Priority Leads - All (Sorted by Score)")

    # Save HIGH priority only
    if high:
        high_xlsx = CAMPAIGNS_DIR / f"PRIORITY_HIGH_{timestamp}.xlsx"
        save_priority_excel(high, high_xlsx, "SURIOTA Priority HIGH - Send First!")

        high_csv = CAMPAIGNS_DIR / f"PRIORITY_HIGH_{timestamp}.csv"
        save_priority_csv(high, high_csv)

    # Save MEDIUM priority
    if medium:
        medium_xlsx = CAMPAIGNS_DIR / f"PRIORITY_MEDIUM_{timestamp}.xlsx"
        save_priority_excel(medium, medium_xlsx, "SURIOTA Priority MEDIUM - Second Wave")

    # Summary
    print("\n" + "=" * 60)
    print("[HIGH PRIORITY - Top 20 Leads to Send First]")
    print("=" * 60)

    for i, lead in enumerate(high[:20], 1):
        print(f"\n{i}. [{lead['priority_score']}] {lead['name']}")
        print(f"   {lead['email']}")
        print(f"   {lead['company']} | {lead['position'] or 'N/A'}")
        print(f"   Reasons: {lead['priority_reasons']}")

    # By category breakdown
    print("\n" + "=" * 60)
    print("[HIGH Priority by Category]")
    print("=" * 60)

    for cat in ["Big Company", "Medium Company", "Micro Company"]:
        cat_high = [c for c in high if c.get("category") == cat]
        print(f"  {cat}: {len(cat_high)} HIGH priority leads")

    # Industry breakdown for HIGH
    print("\n[HIGH Priority by Industry (Top 10)]")
    industries = {}
    for c in high:
        ind = c.get("industry", "Unknown")
        if ind and ind != "industri Anda":
            industries[ind] = industries.get(ind, 0) + 1

    for ind, count in sorted(industries.items(), key=lambda x: -x[1])[:10]:
        print(f"  {ind[:40]}: {count}")

    print("\n" + "=" * 60)
    print("  Done! Priority leads filtered successfully.")
    print(f"  Kirim HIGH priority ({len(high)} leads) terlebih dahulu!")
    print("=" * 60)


if __name__ == "__main__":
    main()
