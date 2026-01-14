"""
Merge All Leads - Big, Medium, Micro Companies
===============================================
Gabungkan semua leads dari berbagai sumber menjadi satu file master
"""

import json
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
TOOLS_DIR = Path(__file__).parent.parent
OUTPUT_DIR = TOOLS_DIR / "output"
LEADS_DIR = OUTPUT_DIR / "leads"
CAMPAIGNS_DIR = OUTPUT_DIR / "campaigns"

# Brand Colors
COLORS = {
    "primary": "1E3A5F",
    "header_text": "FFFFFF",
    "border": "E2E8F0",
    "alt_row": "F1F5F9",
    "big": "2E8B57",      # Green for Big
    "medium": "4169E1",   # Blue for Medium
    "micro": "FF8C00",    # Orange for Micro
}


def load_big_company_leads():
    """Load big company leads from ALL_BATAM_LEADS.json"""
    path = LEADS_DIR / "ALL_BATAM_LEADS.json"
    if not path.exists():
        print(f"[Warning] Not found: {path}")
        return []

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    leads = []
    for d in data:
        leads.append({
            "name": f"{d.get('first_name', '')} {d.get('last_name', '')}".strip() or "Unknown",
            "email": d.get("email", ""),
            "phone": "",
            "company": d.get("company", ""),
            "position": d.get("position", ""),
            "industry": d.get("industry", ""),
            "category": "Big Company",
            "city": "Batam",
            "website": d.get("domain", ""),
            "linkedin": d.get("linkedin", ""),
            "confidence": d.get("confidence", ""),
            "source": "Hunter.io"
        })

    return leads


def load_medium_company_leads():
    """Load medium company leads from MEDIUM_COMPANY_BATAM_COMPRO.json"""
    path = LEADS_DIR / "MEDIUM_COMPANY_BATAM_COMPRO.json"
    if not path.exists():
        print(f"[Warning] Not found: {path}")
        return []

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    leads = []
    for d in data:
        leads.append({
            "name": d.get("name", "Unknown"),
            "email": d.get("email", ""),
            "phone": d.get("phone", ""),
            "company": d.get("company", ""),
            "position": d.get("position", ""),
            "industry": d.get("industry", ""),
            "category": "Medium Company",
            "city": d.get("city", "Batam"),
            "website": d.get("website", ""),
            "linkedin": "",
            "confidence": str(d.get("confidence", "")),
            "source": d.get("source", "Web Scraping")
        })

    return leads


def load_micro_company_leads():
    """Load micro company leads from MICRO_COMPANY_BATAM_COMPRO.json"""
    path = LEADS_DIR / "MICRO_COMPANY_BATAM_COMPRO.json"
    if not path.exists():
        print(f"[Warning] Not found: {path}")
        return []

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    leads = []
    for d in data:
        leads.append({
            "name": d.get("name", "Unknown"),
            "email": d.get("email", ""),
            "phone": d.get("phone", ""),
            "company": d.get("company", ""),
            "position": "",
            "industry": d.get("industry", ""),
            "category": "Micro Company",
            "city": d.get("city", "Batam"),
            "website": "",
            "linkedin": "",
            "confidence": "",
            "source": d.get("source", "Dinas Koperasi Kepri")
        })

    return leads


def deduplicate_leads(leads):
    """Remove duplicate emails"""
    seen_emails = set()
    unique_leads = []

    for lead in leads:
        email = lead.get("email", "").lower().strip()
        if email and email not in seen_emails and "@" in email:
            seen_emails.add(email)
            unique_leads.append(lead)

    return unique_leads


def save_to_json(leads, path):
    """Save leads to JSON file"""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(leads, f, indent=2, ensure_ascii=False)
    print(f"[Saved] {path.name} ({len(leads)} leads)")


def save_to_csv(leads, path):
    """Save leads to CSV file"""
    if not leads:
        return

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=leads[0].keys())
        writer.writeheader()
        writer.writerows(leads)
    print(f"[Saved] {path.name} ({len(leads)} leads)")


def save_to_excel(leads, path, title):
    """Save leads to styled Excel file"""
    if not leads:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "All Leads"

    headers = list(leads[0].keys())

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=COLORS["primary"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Stats subtitle
    big_count = sum(1 for l in leads if l.get("category") == "Big Company")
    medium_count = sum(1 for l in leads if l.get("category") == "Medium Company")
    micro_count = sum(1 for l in leads if l.get("category") == "Micro Company")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    subtitle = f"Total: {len(leads)} leads | Big: {big_count} | Medium: {medium_count} | Micro: {micro_count} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(name="Calibri", size=11, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4

    # Column widths
    column_widths = {
        "name": 25, "email": 35, "phone": 18, "company": 35,
        "position": 25, "industry": 30, "category": 18, "city": 12,
        "website": 25, "linkedin": 40, "confidence": 12, "source": 20
    }

    # Headers
    header_style = {
        "font": Font(name="Calibri", size=11, bold=True, color=COLORS["header_text"]),
        "fill": PatternFill(start_color=COLORS["primary"], end_color=COLORS["primary"], fill_type="solid"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            bottom=Side(style="medium", color=COLORS["primary"]),
            top=Side(style="thin", color=COLORS["border"]),
            left=Side(style="thin", color=COLORS["border"]),
            right=Side(style="thin", color=COLORS["border"])
        )
    }

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header.upper().replace("_", " "))
        cell.font = header_style["font"]
        cell.fill = header_style["fill"]
        cell.alignment = header_style["alignment"]
        cell.border = header_style["border"]

        col_letter = get_column_letter(col)
        width = column_widths.get(header.lower(), 15)
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[header_row].height = 25

    # Data rows
    normal_border = Border(
        bottom=Side(style="thin", color=COLORS["border"]),
        top=Side(style="thin", color=COLORS["border"]),
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"])
    )

    for row_idx, lead in enumerate(leads, header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        category = lead.get("category", "")

        # Category color indicator
        if category == "Big Company":
            cat_color = COLORS["big"]
        elif category == "Medium Company":
            cat_color = COLORS["medium"]
        else:
            cat_color = COLORS["micro"]

        for col_idx, header in enumerate(headers, 1):
            value = lead.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

            cell.font = Font(name="Calibri", size=10, color="333333")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = normal_border

            if is_alt:
                cell.fill = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")

            # Email hyperlink
            if header == "email" and value and "@" in str(value):
                cell.font = Font(name="Calibri", size=10, color="4169E1", underline="single")
                cell.hyperlink = f"mailto:{value}"

            # Category color
            if header == "category":
                cell.font = Font(name="Calibri", size=10, bold=True, color=cat_color)

        ws.row_dimensions[row_idx].height = 20

    # Freeze and filter
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{header_row + len(leads)}"

    wb.save(path)
    print(f"[Saved] {path.name} ({len(leads)} leads)")


def main():
    print("=" * 60)
    print("  SURIOTA All Leads Merger")
    print("  Merging Big + Medium + Micro Companies")
    print("=" * 60)
    print()

    # Load all leads
    print("[Loading leads...]")
    big_leads = load_big_company_leads()
    print(f"  Big Company: {len(big_leads)} leads")

    medium_leads = load_medium_company_leads()
    print(f"  Medium Company: {len(medium_leads)} leads")

    micro_leads = load_micro_company_leads()
    print(f"  Micro Company: {len(micro_leads)} leads")

    # Merge all
    all_leads = big_leads + medium_leads + micro_leads
    print(f"\n[Total before dedup]: {len(all_leads)} leads")

    # Deduplicate
    unique_leads = deduplicate_leads(all_leads)
    print(f"[Total after dedup]: {len(unique_leads)} leads")

    # Sort by category then company
    unique_leads.sort(key=lambda x: (
        0 if x.get("category") == "Big Company" else (1 if x.get("category") == "Medium Company" else 2),
        x.get("company", "").lower()
    ))

    # Save outputs
    print("\n[Saving files...]")
    timestamp = datetime.now().strftime("%Y%m%d")

    # JSON
    json_path = LEADS_DIR / f"ALL_LEADS_MERGED_{timestamp}.json"
    save_to_json(unique_leads, json_path)

    # CSV
    csv_path = LEADS_DIR / f"ALL_LEADS_MERGED_{timestamp}.csv"
    save_to_csv(unique_leads, csv_path)

    # Excel
    excel_path = LEADS_DIR / f"ALL_LEADS_MERGED_{timestamp}.xlsx"
    save_to_excel(unique_leads, excel_path, "SURIOTA All Leads - Big + Medium + Micro Companies")

    # Summary
    print("\n" + "=" * 60)
    print("[Summary]")
    print("=" * 60)

    # By category
    categories = {}
    for lead in unique_leads:
        cat = lead.get("category", "Unknown")
        categories[cat] = categories.get(cat, 0) + 1

    print("\n[By Category]")
    for cat, count in sorted(categories.items()):
        print(f"  {cat}: {count} leads")

    # Company count by category
    print("\n[Companies by Category]")
    for cat in ["Big Company", "Medium Company", "Micro Company"]:
        companies = set(l.get("company", "") for l in unique_leads if l.get("category") == cat)
        print(f"  {cat}: {len(companies)} companies")

    print("\n" + "=" * 60)
    print("  Done! All leads merged successfully.")
    print("=" * 60)


if __name__ == "__main__":
    main()
