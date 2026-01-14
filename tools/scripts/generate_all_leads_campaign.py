"""
Generate Email Campaign for ALL Leads - Big, Medium, Micro Companies
=====================================================================
Target: Semua leads gabungan untuk Company Profile SURIOTA
"""

import json
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paths
TOOLS_DIR = Path(__file__).parent.parent
OUTPUT_DIR = TOOLS_DIR / "output"
LEADS_DIR = OUTPUT_DIR / "leads"
CAMPAIGNS_DIR = OUTPUT_DIR / "campaigns"

# Brand Colors
COLORS = {
    "primary": "1E3A5F",
    "header_text": "FFFFFF",
    "border": "E2E8F0",
    "alt_row": "F1F5F9",
    "big": "2E8B57",
    "medium": "4169E1",
    "micro": "FF8C00",
}

# Email Templates by Category
EMAIL_TEMPLATES = {
    "Big Company": {
        "subject": "Industrial IoT Partnership Opportunity - SURIOTA x {company}",
        "subject_id": "Peluang Kemitraan Industrial IoT - SURIOTA x {company}",
        "body_id": """Kepada Yth. {name},
{position_line}
Salam hangat dari PT Surya Inovasi Prioritas (SURIOTA).

Kami adalah perusahaan teknologi Industrial IoT dan System Integration berbasis di Batam. Melihat {company} sebagai pemimpin industri di sektor {industry}, kami ingin memperkenalkan solusi kami yang dapat mendukung efisiensi operasional perusahaan Anda.

**Solusi Industrial IoT Kami:**

1. **System Integration** - Koneksi PLC/SCADA ke Cloud, Modbus to MQTT
2. **Predictive Maintenance** - Monitoring vibrasi & suhu untuk mencegah downtime
3. **Remote Monitoring Dashboard** - Real-time monitoring via web & mobile
4. **Energy Management System** - Tracking konsumsi listrik & carbon footprint
5. **Asset Tracking** - GPS tracking & fleet management

**Produk Unggulan:**
- **SRT-MGATE-1210** - Industrial Modbus-to-MQTT Gateway (Rp 2.7jt)
- **SURGE Platform** - Multi-tenant IoT Monitoring SaaS (dari $29/bulan)

**Track Record:**
- 55+ proyek sukses di Manufacturing, Energy, Maritime, Logistics
- Klien: Industrial parks, Water utilities, Energy companies

Apakah Bapak/Ibu berkenan untuk meeting singkat (15-20 menit) via Zoom untuk diskusi kebutuhan {company}?

Hormat kami,

Tim SURIOTA
PT Surya Inovasi Prioritas
Website: www.suriota.com
WhatsApp: +62 858-3567-2476
Email: sales@suriota.com
""",
        "body_en": """Dear {name},
{position_line}
Warm greetings from PT Surya Inovasi Prioritas (SURIOTA).

We are an Industrial IoT and System Integration company based in Batam. Recognizing {company} as an industry leader in {industry}, we would like to introduce our solutions that can support your operational efficiency.

**Our Industrial IoT Solutions:**

1. **System Integration** - PLC/SCADA to Cloud, Modbus to MQTT connectivity
2. **Predictive Maintenance** - Vibration & temperature monitoring to prevent downtime
3. **Remote Monitoring Dashboard** - Real-time monitoring via web & mobile
4. **Energy Management System** - Power consumption & carbon footprint tracking
5. **Asset Tracking** - GPS tracking & fleet management

**Key Products:**
- **SRT-MGATE-1210** - Industrial Modbus-to-MQTT Gateway (~$169)
- **SURGE Platform** - Multi-tenant IoT Monitoring SaaS (from $29/month)

**Track Record:**
- 55+ successful projects in Manufacturing, Energy, Maritime, Logistics
- Clients: Industrial parks, Water utilities, Energy companies

Would you be available for a brief meeting (15-20 minutes) via Zoom to discuss {company}'s needs?

Best regards,

SURIOTA Team
PT Surya Inovasi Prioritas
Website: www.suriota.com
WhatsApp: +62 858-3567-2476
Email: sales@suriota.com
"""
    },
    "Medium Company": {
        "subject": "Solusi Industrial IoT untuk {company} - SURIOTA",
        "subject_id": "Solusi Industrial IoT untuk {company} - SURIOTA",
        "body_id": """Kepada Yth. {name},
{position_line}
Salam hangat dari PT Surya Inovasi Prioritas (SURIOTA).

Kami adalah perusahaan teknologi yang bergerak di bidang Industrial IoT Services dan System Integration berbasis di Batam. Melihat {company} sebagai perusahaan terkemuka di bidang {industry}, kami yakin ada peluang kolaborasi yang dapat saling menguntungkan.

**Layanan Kami yang Relevan untuk Industri Anda:**

1. **System Integration** - Koneksi PLC/SCADA ke Cloud, Modbus ke MQTT
2. **Predictive Maintenance** - Monitoring vibrasi & suhu untuk mencegah downtime
3. **Remote Monitoring Dashboard** - Real-time monitoring via web & mobile app
4. **Energy Management System** - Monitoring konsumsi listrik & carbon footprint
5. **Custom IoT Development** - Solusi custom sesuai kebutuhan operasional

**Produk Unggulan:**
- **SRT-MGATE-1210** - Industrial Modbus-to-MQTT Gateway (Rp 2.7jt)
- **SURGE Platform** - Multi-tenant IoT Monitoring SaaS (mulai $29/bulan)

**Track Record:**
- 55+ proyek sukses di sektor Manufacturing, Energy, Maritime, dan Logistics
- Klien: Industrial parks, Water utilities, Energy companies

Apakah Bapak/Ibu berkenan untuk meeting singkat (15-20 menit) via Zoom atau kunjungan ke kantor untuk diskusi lebih lanjut tentang kebutuhan {company}?

Hormat kami,

Tim SURIOTA
PT Surya Inovasi Prioritas
Website: www.suriota.com
WhatsApp: +62 858-3567-2476
Email: sales@suriota.com
""",
        "body_en": """Dear {name},
{position_line}
Warm greetings from PT Surya Inovasi Prioritas (SURIOTA).

We are a technology company specializing in Industrial IoT Services and System Integration based in Batam. Recognizing {company} as a leading company in {industry}, we believe there are mutually beneficial collaboration opportunities.

**Our Services Relevant to Your Industry:**

1. **System Integration** - PLC/SCADA to Cloud, Modbus to MQTT connectivity
2. **Predictive Maintenance** - Vibration & temperature monitoring to prevent downtime
3. **Remote Monitoring Dashboard** - Real-time monitoring via web & mobile app
4. **Energy Management System** - Power consumption & carbon footprint tracking
5. **Custom IoT Development** - Tailored solutions for your operational needs

**Key Products:**
- **SRT-MGATE-1210** - Industrial Modbus-to-MQTT Gateway (~$169)
- **SURGE Platform** - Multi-tenant IoT Monitoring SaaS (from $29/month)

**Track Record:**
- 55+ successful projects in Manufacturing, Energy, Maritime, and Logistics
- Clients: Industrial parks, Water utilities, Energy companies

Would you be available for a brief meeting (15-20 minutes) via Zoom or an office visit to discuss {company}'s specific needs?

Best regards,

SURIOTA Team
PT Surya Inovasi Prioritas
Website: www.suriota.com
WhatsApp: +62 858-3567-2476
Email: sales@suriota.com
"""
    },
    "Micro Company": {
        "subject": "Kolaborasi Bisnis - SURIOTA x {company}",
        "subject_id": "Kolaborasi Bisnis - SURIOTA x {company}",
        "body_id": """Kepada Yth. {name},
{position_line}
Salam hangat dari PT Surya Inovasi Prioritas (SURIOTA).

Kami adalah perusahaan teknologi Industrial IoT dan System Integration berbasis di Batam. Kami melihat potensi kolaborasi dengan {company} di bidang {industry}.

**Tentang SURIOTA:**

Kami menyediakan solusi IoT untuk industri manufaktur, energi, maritim, dan logistik:
- System Integration (PLC/SCADA to Cloud)
- Remote Monitoring Dashboard
- Asset Tracking & Management
- Custom IoT Development

**Produk Kami:**
- **SRT-MGATE-1210** - Industrial Gateway (Rp 2.7jt)
- **SURGE Platform** - IoT Monitoring SaaS (dari $29/bulan)

**Peluang Kolaborasi:**
- Kemitraan teknologi
- Referral program
- Project collaboration

Apakah Bapak/Ibu berkenan untuk diskusi singkat tentang potensi kolaborasi?

Hormat kami,

Tim SURIOTA
PT Surya Inovasi Prioritas
Website: www.suriota.com
WhatsApp: +62 858-3567-2476
Email: sales@suriota.com
""",
        "body_en": """Dear {name},
{position_line}
Warm greetings from PT Surya Inovasi Prioritas (SURIOTA).

We are an Industrial IoT and System Integration company based in Batam. We see collaboration potential with {company} in {industry}.

**About SURIOTA:**

We provide IoT solutions for manufacturing, energy, maritime, and logistics industries:
- System Integration (PLC/SCADA to Cloud)
- Remote Monitoring Dashboard
- Asset Tracking & Management
- Custom IoT Development

**Our Products:**
- **SRT-MGATE-1210** - Industrial Gateway (~$169)
- **SURGE Platform** - IoT Monitoring SaaS (from $29/month)

**Collaboration Opportunities:**
- Technology partnership
- Referral program
- Project collaboration

Would you be available for a brief discussion about collaboration potential?

Best regards,

SURIOTA Team
PT Surya Inovasi Prioritas
Website: www.suriota.com
WhatsApp: +62 858-3567-2476
Email: sales@suriota.com
"""
    }
}


def load_merged_leads():
    """Load merged leads from today's file"""
    timestamp = datetime.now().strftime("%Y%m%d")
    path = LEADS_DIR / f"ALL_LEADS_MERGED_{timestamp}.json"

    if not path.exists():
        # Try to find any merged file
        merged_files = list(LEADS_DIR.glob("ALL_LEADS_MERGED_*.json"))
        if merged_files:
            path = sorted(merged_files)[-1]  # Get latest
        else:
            print(f"[Error] No merged leads file found")
            return []

    with open(path, "r", encoding="utf-8") as f:
        leads = json.load(f)

    print(f"[Loaded] {len(leads)} leads from {path.name}")
    return leads


def generate_campaign(leads):
    """Generate email campaign for all leads"""
    campaign_rows = []
    seen_emails = set()

    skip_positions = ["Sales", "Contact", "Admin", "Support", "Contact Person",
                      "Sales Team", "Info Team", "Admin Team", "General"]

    for lead in leads:
        email = lead.get("email", "").strip().lower()
        if not email or "@" not in email:
            continue
        if email in seen_emails:
            continue
        seen_emails.add(email)

        name = lead.get("name", "Bapak/Ibu")
        if name == "Unknown" or not name:
            name = "Bapak/Ibu"

        company = lead.get("company", "perusahaan Anda")
        if not company:
            company = "perusahaan Anda"

        industry = lead.get("industry", "industri Anda")
        if not industry:
            industry = "industri Anda"

        position = lead.get("position", "")
        category = lead.get("category", "Medium Company")

        # Create position line if available
        position_line = ""
        if position and position not in skip_positions:
            position_line = f"({position})\n"

        # Get template based on category
        template = EMAIL_TEMPLATES.get(category, EMAIL_TEMPLATES["Medium Company"])

        # Generate personalized email
        subject = template["subject"].format(company=company)
        body_id = template["body_id"].format(
            name=name,
            company=company,
            industry=industry,
            position_line=position_line
        )
        body_en = template["body_en"].format(
            name=name,
            company=company,
            industry=industry,
            position_line=position_line
        )

        campaign_rows.append({
            "email": email,
            "name": name,
            "company": company,
            "position": position,
            "industry": industry,
            "phone": lead.get("phone", ""),
            "category": category,
            "city": lead.get("city", "Batam"),
            "subject": subject,
            "body_id": body_id,
            "body_en": body_en,
            "confidence": lead.get("confidence", ""),
            "source": lead.get("source", "")
        })

    return campaign_rows


def save_to_csv(campaigns, path):
    """Save campaigns to CSV"""
    if not campaigns:
        return

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=campaigns[0].keys())
        writer.writeheader()
        writer.writerows(campaigns)

    print(f"[Saved] {path.name} ({len(campaigns)} campaigns)")


def save_to_json(campaigns, path):
    """Save campaigns to JSON"""
    with open(path, "w", encoding="utf-8") as f:
        json.dump(campaigns, f, indent=2, ensure_ascii=False)

    print(f"[Saved] {path.name} ({len(campaigns)} campaigns)")


def save_to_excel(campaigns, path, title):
    """Save campaigns to styled Excel"""
    if not campaigns:
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Email Campaign"

    # Only include key columns for Excel (body too long)
    excel_columns = ["email", "name", "company", "position", "industry",
                     "phone", "category", "city", "subject", "confidence", "source"]

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(excel_columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", size=16, bold=True, color=COLORS["primary"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 30

    # Stats
    big_count = sum(1 for c in campaigns if c.get("category") == "Big Company")
    medium_count = sum(1 for c in campaigns if c.get("category") == "Medium Company")
    micro_count = sum(1 for c in campaigns if c.get("category") == "Micro Company")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(excel_columns))
    subtitle = f"Total: {len(campaigns)} emails | Big: {big_count} | Medium: {medium_count} | Micro: {micro_count} | Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    subtitle_cell = ws.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(name="Calibri", size=11, italic=True, color="666666")
    subtitle_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 20

    header_row = 4

    # Column widths
    column_widths = {
        "email": 35, "name": 20, "company": 35, "position": 20,
        "industry": 25, "phone": 18, "category": 15, "city": 12,
        "subject": 45, "confidence": 12, "source": 15
    }

    # Header style
    header_fill = PatternFill(start_color=COLORS["primary"], end_color=COLORS["primary"], fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color=COLORS["header_text"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        bottom=Side(style="thin", color=COLORS["border"]),
        top=Side(style="thin", color=COLORS["border"]),
        left=Side(style="thin", color=COLORS["border"]),
        right=Side(style="thin", color=COLORS["border"])
    )

    # Headers
    for col, header in enumerate(excel_columns, 1):
        cell = ws.cell(row=header_row, column=col, value=header.upper().replace("_", " "))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = column_widths.get(header, 15)

    ws.row_dimensions[header_row].height = 25

    # Data rows
    for row_idx, campaign in enumerate(campaigns, header_row + 1):
        is_alt = (row_idx - header_row) % 2 == 0
        category = campaign.get("category", "")

        # Category color
        if category == "Big Company":
            cat_color = COLORS["big"]
        elif category == "Medium Company":
            cat_color = COLORS["medium"]
        else:
            cat_color = COLORS["micro"]

        for col_idx, header in enumerate(excel_columns, 1):
            value = campaign.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")

            cell.font = Font(name="Calibri", size=10, color="333333")
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = border

            if is_alt:
                cell.fill = PatternFill(start_color=COLORS["alt_row"], end_color=COLORS["alt_row"], fill_type="solid")

            # Email hyperlink
            if header == "email" and value and "@" in str(value):
                cell.font = Font(name="Calibri", size=10, color="4169E1", underline="single")
                cell.hyperlink = f"mailto:{value}"

            # Category color
            if header == "category":
                cell.font = Font(name="Calibri", size=10, bold=True, color=cat_color)

        ws.row_dimensions[row_idx].height = 20

    # Freeze and filter
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(excel_columns))}{header_row + len(campaigns)}"

    wb.save(path)
    print(f"[Saved] {path.name} ({len(campaigns)} campaigns)")


def main():
    print("=" * 60)
    print("  SURIOTA Email Campaign Generator")
    print("  All Leads - Big + Medium + Micro Companies")
    print("=" * 60)
    print()

    # Load merged leads
    leads = load_merged_leads()
    if not leads:
        print("[Error] No leads to process")
        return

    # Generate campaigns
    print("\n[Generating email campaigns...]")
    campaigns = generate_campaign(leads)
    print(f"[Generated] {len(campaigns)} email campaigns")

    # Save outputs
    print("\n[Saving files...]")
    timestamp = datetime.now().strftime("%Y%m%d")

    # CSV
    csv_path = CAMPAIGNS_DIR / f"EMAIL_CAMPAIGN_ALL_LEADS_{timestamp}.csv"
    save_to_csv(campaigns, csv_path)

    # JSON
    json_path = CAMPAIGNS_DIR / f"EMAIL_CAMPAIGN_ALL_LEADS_{timestamp}.json"
    save_to_json(campaigns, json_path)

    # Excel
    excel_path = CAMPAIGNS_DIR / f"EMAIL_CAMPAIGN_ALL_LEADS_{timestamp}.xlsx"
    save_to_excel(campaigns, excel_path, "SURIOTA Email Campaign - All Leads (Big + Medium + Micro)")

    # Summary
    print("\n" + "=" * 60)
    print("[Summary]")
    print("=" * 60)

    # By category
    categories = {}
    for c in campaigns:
        cat = c.get("category", "Unknown")
        categories[cat] = categories.get(cat, 0) + 1

    print("\n[By Category]")
    for cat in ["Big Company", "Medium Company", "Micro Company"]:
        count = categories.get(cat, 0)
        print(f"  {cat}: {count} emails")

    # By industry (top 10)
    industries = {}
    for c in campaigns:
        ind = c.get("industry", "Unknown")
        if ind and ind != "industri Anda":
            industries[ind] = industries.get(ind, 0) + 1

    print("\n[Top 10 Industries]")
    for ind, count in sorted(industries.items(), key=lambda x: -x[1])[:10]:
        print(f"  {ind}: {count}")

    # By company (top 10)
    companies = {}
    for c in campaigns:
        comp = c.get("company", "Unknown")
        if comp and comp != "perusahaan Anda":
            companies[comp] = companies.get(comp, 0) + 1

    print("\n[Top 10 Companies by Contacts]")
    for comp, count in sorted(companies.items(), key=lambda x: -x[1])[:10]:
        print(f"  {comp}: {count}")

    # Preview
    print("\n" + "=" * 60)
    print("[Preview - Sample Emails by Category]")
    print("=" * 60)

    for cat in ["Big Company", "Medium Company", "Micro Company"]:
        sample = next((c for c in campaigns if c.get("category") == cat), None)
        if sample:
            print(f"\n--- {cat} ---")
            print(f"To: {sample['email']}")
            print(f"Subject: {sample['subject']}")
            print(f"Company: {sample['company']}")

    print("\n" + "=" * 60)
    print("  Done! Email campaigns generated successfully.")
    print("=" * 60)


if __name__ == "__main__":
    main()
